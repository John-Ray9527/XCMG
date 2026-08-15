# -*- coding: utf-8 -*-
"""生成 CAT（卡特彼勒）90-200 吨级液压挖掘机产品清单 xlsx"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
OUT = os.path.join(DESKTOP, "CAT_90-200吨_液压挖掘机产品清单.xlsx")

HEADERS = [
    "型号", "产品系列", "结构形式", "工作重量(t)",
    "发动机", "额定功率(kW)", "功率(hp)", "斗容(m³)", "额定载荷(t)",
    "液压/技术要点", "备注",
]

# 现役 90~200 吨机型
ROWS = [
    ["395", "大型液压挖掘机", "反铲", "94~95",
     "Cat C18", "404~405", "542", "5.7~8.5（标准6.5）", "—",
     "变量泵 + 负载敏感控制；斗杆挖掘力497 kN / 铲斗394 kN；回转速度6.26 rpm",
     "90吨级现役主力；中国区 NR III 版约95 t / 568 kW"],
    ["6015", "液压矿用挖掘机", "正铲/反铲", "140",
     "Cat C27", "615", "824", "8.1（1:1）", "14.6",
     "3×变量斜盘泵，最大流量1544 L/min，附件压力320 bar；按需先导泵节油≥5.5%",
     "匹配 Cat 773/775/777 矿卡（55~100吨级）"],
    ["6018 / 6018 FS", "液压矿用挖掘机", "正铲/反铲", "183",
     "2× Cat C18 ACERT", "858", "1150", "10.0（SAE 2:1）", "—",
     "双发动机配置；标准斗10 m³，可选11.4 m³",
     "整备质量约177 t（正铲）/181 t（反铲），带料约195 t"],
]

# 边界 / 上代机型
BOUNDARY = [
    ["390F L", "大型液压挖掘机", "反铲", "~87（低于90）",
     "Cat C18 ACERT", "391", "524", "3.9~6.0", "—",
     "上代90吨级，已被395取代", "边界机型，官方整机86.8 t"],
]

NOTES = [
    "1. 数据来源：Caterpillar 官网 / 经销商规格页 / 官方产品手册，检索日期 2026-08。",
    "2. 工作重量随臂长、斗杆、铲斗、履带板等配置不同而变化，表中为典型配置值。",
    "3. 「90~200吨」指整机工作重量区间；390F 为边界机型（约87 t，略低于90 t），一并列出供参考。",
    "4. 已停产但在役机型：CAT 385C（约91.6 t）、5090B（约91.6 t），未列入主表。",
    "5. 6018 采用双 Cat C18 发动机，属该吨位级少见的双动力配置，值得竞品对标关注。",
]

wb = Workbook()
ws = wb.active
ws.title = "CAT_90-200吨产品"

# 样式
title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F2937")
head_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
head_fill = PatternFill("solid", fgColor="C99700")  # 卡特黄
body_font = Font(name="微软雅黑", size=10, color="1F2937")
bound_font = Font(name="微软雅黑", size=10, color="6B7280", italic=True)
note_font = Font(name="微软雅黑", size=9, color="6B7280")
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ncol = len(HEADERS)

# 标题行
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
c = ws.cell(row=1, column=1, value="CATERPILLAR（卡特彼勒）90–200 吨级液压挖掘机产品清单")
c.font = title_font
c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# 表头
for j, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=j, value=h)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[2].height = 24

# 数据行
def write_rows(start, rows, font):
    r = start
    for row in rows:
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = font
            cell.border = border
            cell.alignment = center if j <= 4 else wrap
        r += 1
    return r

r = write_rows(3, ROWS, body_font)

# 边界机型分隔
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
c = ws.cell(row=r, column=1, value="边界 / 上代机型")
c.font = Font(name="微软雅黑", size=10, bold=True, color="1F2937")
c.fill = PatternFill("solid", fgColor="F3F4F6")
c.alignment = Alignment(horizontal="left", vertical="center")
r += 1
r = write_rows(r, BOUNDARY, bound_font)

# 说明
r += 1
for note in NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    cell = ws.cell(row=r, column=1, value=note)
    cell.font = note_font
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    r += 1

# 列宽
widths = [18, 18, 12, 14, 20, 14, 10, 20, 12, 46, 34]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w

# 冻结表头
ws.freeze_panes = "A3"

# 行高（数据行自适应，设一个基础高度）
for row in range(3, 3 + len(ROWS) + len(BOUNDARY) + 1):
    if row not in (r - len(NOTES) - 1,):  # 简单起见统一设
        ws.row_dimensions[row].height = 44

wb.save(OUT)
print("已生成:", OUT)
